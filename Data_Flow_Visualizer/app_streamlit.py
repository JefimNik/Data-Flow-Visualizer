import streamlit as st
import pandas as pd
import yaml
import pathlib
import subprocess
import os
from io import BytesIO
import openpyxl

# --------- Пути ---------
BASE = pathlib.Path(__file__).resolve().parent
CONFIG_PATH = BASE / "config" / "data_model.yaml"
GENERATOR = BASE / "src" / "generate_html.py"
BUILD_HTML = BASE / "build" / "data_model_v1.html"

# --------- Работа с YAML ---------
def load_yaml(path):
    with open(path, "r", encoding="utf-8") as f:
        return yaml.safe_load(f)

def save_yaml(path, data):
    with open(path, "w", encoding="utf-8") as f:
        yaml.safe_dump(data, f, allow_unicode=True, sort_keys=False)
        f.flush()
        os.fsync(f.fileno())

# --------- Генерация HTML ---------
def generate_html(selected_node=None):
    """Генерирует HTML и передаёт ID выделенного узла для подсветки."""
    try:
        result = subprocess.run(
            ["python", str(GENERATOR)],
            capture_output=True,
            text=True
        )
        if result.returncode != 0:
            st.error("❌ Ошибка при генерации HTML:")
            st.code(result.stderr or result.stdout)
        else:
            st.success("✅ HTML успешно сгенерирован.")
    except Exception as e:
        st.error(f"⚠️ Не удалось запустить генерацию: {e}")

# --------- Создание Excel-шаблона узла ---------
def make_excel(node: dict) -> BytesIO:
    """Создаёт Excel-файл с данными узла и таблицей колонок, начиная с 7-й строки."""
    buffer = BytesIO()
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        sheet_name = "node"
        pd.DataFrame().to_excel(writer, index=False, sheet_name=sheet_name)
        ws = writer.sheets[sheet_name]

        ws["A1"] = "Узел данных"
        ws["A2"] = "name"; ws["B2"] = node.get("name", "")
        ws["A3"] = "layer"; ws["B3"] = node.get("layer", "")
        ws["A4"] = "type"; ws["B4"] = node.get("type", "")
        ws["A5"] = "comment"; ws["B5"] = node.get("comment", "")

        ws["A7"] = "Таблица колонок:"
        headers = ["name", "type", "description", "comment"]
        for i, h in enumerate(headers, start=1):
            ws.cell(row=8, column=i, value=h)

        cols = pd.DataFrame(node.get("columns", []))
        if cols.empty:
            cols = pd.DataFrame(columns=headers)

        for r, row in enumerate(cols.itertuples(index=False), start=9):
            for c, value in enumerate(row, start=1):
                ws.cell(row=r, column=c, value=value)

    buffer.seek(0)
    return buffer

# --------- Санитизация таблицы (во избежание Arrow ошибок) ---------
def sanitize_dataframe(df: pd.DataFrame) -> pd.DataFrame:
    if df.empty:
        return pd.DataFrame(columns=["name", "type", "description", "comment"])
    df = df.fillna("").astype(str)
    df.columns = df.columns.map(str)
    return df

# --------- Интерфейс Streamlit ---------
st.set_page_config(page_title="Data Flow Visualizer Editor", layout="wide")
st.title("🧩 Data Flow Visualizer — YAML + Excel Editor")

# ---------- ВЕРХНИЙ БЛОК ----------
st.subheader("⚙️ Экспорт и обновление узлов через Excel")

if not CONFIG_PATH.exists():
    st.error(f"❌ Файл YAML не найден: {CONFIG_PATH}")
    st.stop()

data_model = load_yaml(CONFIG_PATH)
if not data_model or "nodes" not in data_model:
    st.error("❌ Некорректный YAML (нет ключа 'nodes').")
    st.stop()

node_names = [n.get("name", "без имени") for n in data_model.get("nodes", [])]
selected_node_name = st.selectbox("Выберите узел для экспорта", node_names)
node = next((n for n in data_model["nodes"] if n.get("name") == selected_node_name), None)

if node:
    st.markdown(f"#### 🧱 Узел: `{node['name']}`")

    # 📥 Скачать Excel-шаблон
    st.download_button(
        label="📥 Скачать Excel шаблон с данными узла (всё на одном листе)",
        data=make_excel(node),
        file_name=f"{selected_node_name}_template.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )

    # 📤 Загрузить изменённый Excel
    st.markdown("##### 📤 Загрузите изменённый Excel-файл узла")
    uploaded = st.file_uploader("Выберите Excel-файл", type=["xlsx"], key="upload_excel")

    if uploaded is not None:
        try:
            wb = openpyxl.load_workbook(uploaded)
            ws = wb.active

            # Читаем базовые параметры
            name = ws["B2"].value or ""
            layer = ws["B3"].value or ""
            type_ = ws["B4"].value or ""
            comment = ws["B5"].value or ""

            # Читаем таблицу колонок
            start_row = 9
            headers = [ws.cell(row=8, column=i).value for i in range(1, 5)]
            data_rows = []
            for r in range(start_row, ws.max_row + 1):
                row_data = {headers[i - 1]: ws.cell(row=r, column=i).value for i in range(1, len(headers) + 1)}
                if any(v is not None for v in row_data.values()):
                    data_rows.append(row_data)
            cols_df = sanitize_dataframe(pd.DataFrame(data_rows))

            # Обновляем YAML
            new_name = str(name).strip() or selected_node_name
            updated_node = {
                "name": new_name,
                "layer": str(layer).strip(),
                "type": str(type_).strip(),
                "comment": str(comment).strip(),
                "columns": cols_df.to_dict(orient="records"),
            }

            existing_names = [n["name"] for n in data_model["nodes"]]
            if new_name != selected_node_name and new_name in existing_names:
                st.error(f"❌ Узел с именем '{new_name}' уже существует.")
                st.stop()

            replaced = False
            for i, n in enumerate(data_model["nodes"]):
                if n.get("name") == selected_node_name:
                    data_model["nodes"][i] = updated_node
                    replaced = True
                    break
            if not replaced:
                data_model["nodes"].append(updated_node)

            # 💾 Сохраняем YAML и сразу перегенерируем HTML
            save_yaml(CONFIG_PATH, data_model)
            generate_html(selected_node_name)

            st.success(f"✅ YAML обновлён. Узел '{selected_node_name}' → '{new_name}'. HTML перегенерирован.")
            st.dataframe(cols_df)

            # Обновляем страницу с подсветкой выбранного узла
            st.markdown(
                f"<script>window.location.reload(); sessionStorage.setItem('highlightNode', '{new_name}');</script>",
                unsafe_allow_html=True,
            )

        except Exception as e:
            st.error(f"Ошибка при обработке Excel: {e}")

# ---------- НИЖНИЙ БЛОК ----------
st.markdown("---")
st.subheader("🔗 Визуализация модели данных")

try:
    if BUILD_HTML.exists():
        html_code = BUILD_HTML.read_text(encoding="utf-8")
        js_script = """
        <script>
        const lastNode = sessionStorage.getItem('highlightNode');
        if (lastNode && typeof highlightNode === 'function') {
            setTimeout(() => highlightNode(lastNode), 1000);
            sessionStorage.removeItem('highlightNode');
        }
        </script>
        """
        st.components.v1.html(html_code + js_script, height=850, scrolling=True)
    else:
        st.warning("⚠️ HTML не найден. Сначала сгенерируйте его через Excel.")
except Exception as e:
    st.error(f"Ошибка отображения HTML: {e}")
